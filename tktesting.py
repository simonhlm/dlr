# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, datetime, re, os
import chardet 

FIELDS = ["report_level", "country_or_company", "start_date", "end_date", \
          "report_format", "report_criteria", "report_output_level", \
          "account_id", "department", "serial_number", "workitem" ]

class DownloadReport(object):
    def __init__(self, username, password, **kw):

        self.username = username
        self.password = password
        self.record = kw

        """
        profile = webdriver.FirefoxProfile()
        profile.set_preference('browser.download.dir', 'd:\\')
        profile.set_preference('browser.download.folderList', 2)
        profile.set_preference('browser.download.manager.showWhenStarting', False)
        profile.set_preference('browser.helperApps.neverAsk.saveToDisk', 'application/zip')
        driver = webdriver.Firefox(firefox_profile=profile)
        """

    def start(self):


        print(type(self.record["report_level"]),record["report_level"])
        for key in self.record:
            print(key, self.record[key])

        print('%s-%s-%s' % (self.record['start_date'].year, self.record['start_date'].month, self.record['start_date'].day))
        from_date = '-'.join([str(self.record['start_date'].year), str(self.record['start_date'].month), str(self.record['start_date'].day)])
        print(from_date)
        print(self.record["report_level"].encode('utf-8'), 'Country Level'.encode('utf-8'))

        #print(chardet.detect(self.record["report_level"]),"what", chardet.detect("Country"))
        #print(isinstance(self.record["report_level"],unicode))
        if self.record["report_level"] == "Country Level":
            print('report level ok')
        else:
            pass 

        #check for select country or company 
        if self.record["country_or_company"] is 'Australia':
            print('company ok')
        else:
            pass

        #check for report criteria 
        if self.record["report_criteria"] is 'Account ID':
            print('account ok')
        else:
            pass


    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()

def readxls(xlsname):
    try:
        wb = load_workbook(xlsname)
        sheet = wb.get_sheet_by_name("input")
        sheet_parameter = wb.get_sheet_by_name("parameter")
    except Exception as e:
        raise

    start_row = 2
    sheet_rows = []

    for rows in range(start_row, sheet.max_row+1):
        single_row = []
        for cells in sheet[rows]:
            single_row.append(cells.value)
        sheet_rows.append(dict(zip(FIELDS, single_row)))


    #process parameters 
    # check for select report level in column A & B
    report_level_value = []
    report_level_key = []
    for cell in sheet_parameter['A']:
        if cell.value is not None:
            report_level_value.append(cell.value)

    for cell in sheet_parameter['B']:
        if cell.value is not None:
            report_level_key.append(cell.value)

    report_level_dict = dict(zip(report_level_key, report_level_value))

    for key in report_level_dict:
        print(key, report_level_dict[key])


    #check for select country or company 
    country_or_company_value = []
    country_or_company_key = []
    for cell in sheet_parameter['C']:
        if cell.value is not None:
            country_or_company_value.append(cell.value)

    for cell in sheet_parameter['D']:
        if cell.value is not None:
            country_or_company_key.append(cell.value)

    country_or_company_dict = dict(zip(country_or_company_key, country_or_company_value))


    for key in country_or_company_dict:
        print(key, country_or_company_dict[key])


    #check for select report format
    report_format_value = []
    report_format_key = []
    for cell in sheet_parameter['E']:
        if cell.value is not None:
            report_format_value.append(cell.value)

    for cell in sheet_parameter['F']:
        if cell.value is not None:
            report_format_key.append(cell.value)

    report_format_dict = dict(zip(report_format_key, report_format_value))

    for key in report_format_dict:
        print(key, report_format_dict[key])

    #check for report criteria 
    report_criteria_value = []
    report_criteria_key = []
    for cell in sheet_parameter['G']:
        if cell.value is not None:
            report_criteria_value.append(cell.value)

    for cell in sheet_parameter['H']:
        if cell.value is not None:
            report_criteria_key.append(cell.value)

    report_criteria_dict = dict(zip(report_criteria_key, report_criteria_value))

    for key in report_criteria_dict:
        print(key, report_criteria_dict[key])


    #check for report criteria 
    report_output_level_value = []
    report_output_level_key = []
    for cell in sheet_parameter['I']:
        if cell.value is not None:
            report_output_level_value.append(cell.value)

    for cell in sheet_parameter['J']:
        if cell.value is not None:
            report_output_level_key.append(cell.value)

    report_output_level_dict = dict(zip(report_output_level_key, report_output_level_value))

    for key in report_output_level_dict:
        print(key, report_output_level_dict[key])

    return sheet_rows

if __name__ == "__main__":
    # predefine username & password for testing purpose
    records = readxls('dlr.xlsx')
    username = "huanglmw@cn.ibm.com"
    password = "huang005"

    for record in records:
        download_instance = DownloadReport(username, password, **record)
        try:
            download_instance.start()
        except Exception as e:
            print(e)