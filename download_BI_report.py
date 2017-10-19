# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from getpass import getpass

import time, re, os, logging, msvcrt

logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s %(filename)s[line:%(lineno)d]\
                     %(levelname)s %(message)s',
                    datefmt='%a, %d %b %Y %H:%M:%S', 
                    filename='app.log',
                    filemode='w'
                    )

FIELDS = ["report_level", "country_or_company", "start_date", "end_date", \
          "report_format", "report_criteria", "report_output_level", \
          "account_id", "department", "serial_number", "workitem" ]

max_retry_times = 3

labor_report_url = "w3.ibm.com/services/bicentral/processor.wss?\
action=3&newPage=23&FamilySelection=IMGLABVA&sid=170"

class LogonFailedException(BaseException):
    pass

class DownloadReport(object):
    def __init__(self, username, password, **kw):

        self.options = webdriver.ChromeOptions()
        self.prefs = {'profile.default_content_settings.popups': 0, \
        'download.default_directory': os.getcwd()}
        self.options.add_experimental_option('prefs', self.prefs)
        
        """
        fp = webdriver.FirefoxProfile()

        fp.set_preference("browser.download.folderList",2)
        fp.set_preference("browser.download.manager.showWhenStarting",False)
        fp.set_preference("browser.download.dir", os.getcwd())
        fp.set_preference("browser.helperApps.neverAsk.saveToDisk", \
        "application/octet-stream")

        self.driver = webdriver.Firefox(firefox_profile=fp)
        """

        self.username = username
        self.password = password
        self.record = kw

    
    def start(self):
        self.base_url = "https://"+self.username+":"+self.password+"@"+labor_report_url
        self.driver = webdriver.Chrome(chrome_options=self.options)
        self.driver.implicitly_wait(15)
        driver = self.driver

    def verify_windows_security(self):      
        # check for windows security 
        try:
            self.driver.get(self.base_url)
            driver = self.driver
            element = WebDriverWait(self.driver, 3).until(lambda driver: \
                driver.find_element_by_link_text("SIM Various Labor Report"))
        except Exception as e:
            logging.error('Logon failed'+e)
            raise LogonFailedException

        # click the labor report botton 
    def run_labor_report(self):
        
        self.driver.find_element_by_link_text("SIM Various Labor Report").click()

        #check for the logon 
        nowhandle=self.driver.current_window_handle
        allhandles=self.driver.window_handles

        for handle in allhandles:
            if handle != nowhandle:
                self.driver.switch_to.window(handle)
                self.driver.find_element_by_id("CAMUsername").clear()
                self.driver.find_element_by_id("CAMUsername").send_keys(self.username)
                self.driver.find_element_by_id("CAMPassword").clear()
                self.driver.find_element_by_id("CAMPassword").send_keys(self.password)
                self.driver.find_element_by_id("cmdOK").click()
                time.sleep(3)

                # check for select report level
                self.driver.find_element_by_css_selector("option[value=\""+\
                    report_level_dict[self.record["report_level"]]+"\"]").click()

                #check for select country or company 
                self.driver.find_element_by_css_selector("option[value=\""+\
                    country_or_company_dict[self.record["country_or_company"]]+"\"]").click()

                #check for input date 
                input_date = self.driver.find_elements_by_css_selector("input[id^=\"txtDateN\"]")
                from_date = '-'.join([str(self.record['start_date'].year), \
                    str(self.record['start_date'].month), str(self.record['start_date'].day)])
                input_date[0].clear()
                input_date[0].send_keys(from_date)

                to_date = '-'.join([str(self.record['end_date'].year), \
                    str(self.record['end_date'].month), str(self.record['end_date'].day)])
                input_date[1].clear()
                input_date[1].send_keys(to_date)

                #check for select report format
                self.driver.find_element_by_css_selector("option[value=\""+\
                    report_format_dict[self.record["report_format"]]+"\"]").click()
                time.sleep(1)

                #check for report criteria 
                self.driver.find_element_by_css_selector("option[value=\""+\
                    report_criteria_dict[self.record["report_criteria"]]+"\"]").click()

                #check for employee or account id
                self.driver.find_element_by_css_selector("input[value=\""+\
                    report_output_level_dict[self.record["report_output_level"]]+"\"]").click()

                if self.record["report_output_level"] == 'Account':
                    self.driver.find_element_by_css_selector("input[id ^=\"PRMT_TB_\"]").\
                    send_keys(self.record["account_id"])
                elif self.record["report_output_level"] == 'Employee':
                    self.driver.find_element_by_css_selector("input[id ^=\"PRMT_TB_\"]").\
                    send_keys(self.record["serial_number"])
                else:
                    raise

                self.driver.find_element_by_css_selector("button[id ^=\"finishN\"]").click()
                time.sleep(5)
                self.driver.find_element_by_css_selector("span[lid=\"HLExportReportResult_NS_\"]").click()
                time.sleep(10)
                print('The file has download to :'+os.getcwd())
                self.driver.quit()

    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def tearDown(self):
        self.driver.quit()

def loadxls(xlsname):

    try:
        wb = load_workbook(xlsname)
        sheet = wb.get_sheet_by_name("input")
        sheet_parameter = wb.get_sheet_by_name("parameter")
    except Exception as e:
        raise

    start_row = 2
    sheet_rows = []

    for rows in range(start_row, sheet.max_row+1):
        single_row = []
        for cells in sheet[rows]:
            single_row.append(cells.value)
        sheet_rows.append(dict(zip(FIELDS, single_row)))

    #process parameters 
    # check for select report level in column A & B
    global report_level_dict
    report_level_value = []
    report_level_key = []

    for cell in sheet_parameter['A']:
        if cell.value is not None:
            report_level_value.append(cell.value)

    for cell in sheet_parameter['B']:
        if cell.value is not None:
            report_level_key.append(cell.value)

    report_level_dict = dict(zip(report_level_key, report_level_value))

    #check for select country or company 
    global country_or_company_dict 
    country_or_company_value = []
    country_or_company_key = []
    for cell in sheet_parameter['C']:
        if cell.value is not None:
            country_or_company_value.append(cell.value)

    for cell in sheet_parameter['D']:
        if cell.value is not None:
            country_or_company_key.append(cell.value)

    country_or_company_dict = dict(zip(country_or_company_key, country_or_company_value))

    #check for select report format
    global report_format_dict
    report_format_value = []
    report_format_key = []
    for cell in sheet_parameter['E']:
        if cell.value is not None:
            report_format_value.append(cell.value)

    for cell in sheet_parameter['F']:
        if cell.value is not None:
            report_format_key.append(cell.value)

    report_format_dict = dict(zip(report_format_key, report_format_value))

    #check for report criteria 
    global report_criteria_dict
    report_criteria_value = []
    report_criteria_key = []
    for cell in sheet_parameter['G']:
        if cell.value is not None:
            report_criteria_value.append(cell.value)

    for cell in sheet_parameter['H']:
        if cell.value is not None:
            report_criteria_key.append(cell.value)

    report_criteria_dict = dict(zip(report_criteria_key, report_criteria_value))

    #check for report criteria 
    global report_output_level_dict
    report_output_level_value = []
    report_output_level_key = []
    for cell in sheet_parameter['I']:
        if cell.value is not None:
            report_output_level_value.append(cell.value)

    for cell in sheet_parameter['J']:
        if cell.value is not None:
            report_output_level_key.append(cell.value)

    report_output_level_dict = dict(zip(report_output_level_key, report_output_level_value))

    return sheet_rows

def login():
    username = input("Please input your username: ")
    password = getpass("Please entry your password:")

    return username, password

def show_end():
    print('Down Load report have completed, press any key to exit')
    if msvcrt.getch():
        pass

if __name__ == "__main__":
    # predefine username & password for testing purpose
    logging.info('Start to load excel')
    records = loadxls('dlr.xlsx')
    
    logging.info('Get the username and password')
    username, password = login()
    
    logging.info('Start to process request')
    for record in records:
        logon_check = True
        while logon_check and max_retry_times > 0:
            try:
                download_instance = DownloadReport(username, password, **record)
                download_instance.start()
                download_instance.verify_windows_security()
                download_instance.run_labor_report()
                logon_check = False
            except LogonFailedException as e:
                download_instance.tearDown()
                max_retry_times -= 1
                logging.info('Logon failed, please try again')
                print('Logon failed, please try again')
                username, password = login()
            except Exception as e:
                max_retry_times -= 1
                logging.info('Ooops! Error... retry to process the record')
                print('Ooops! Error... retry to process the record')
                download_instance.tearDown()
                logging.error(e)
    show_end() 