# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from getpass import getpass

import time, re, os

FIELDS = ["report_level", "country_or_company", "start_date", "end_date", \
          "report_format", "report_criteria", "report_output_level", \
          "account_id", "department", "serial_number", "workitem" ]

class DownloadReport(object):
    def __init__(self, username, password, **kw):

        self.options = webdriver.ChromeOptions()
        self.prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': os.getcwd()}
        self.options.add_experimental_option('prefs', self.prefs)
        self.driver = webdriver.Chrome(chrome_options=self.options)
        self.username = username
        self.password = password
        self.record = kw

        self.driver.implicitly_wait(30)
        self.base_url = "https://"+self.username+":"+self.password+"@w3.ibm.com/services/bicentral/processor.wss?action=3&newPage=23&FamilySelection=IMGLABVA&sid=170"
    
    def start(self):
        driver = self.driver
        try:
            driver.get(self.base_url)
        except Exception as e:
            raise
        
        driver.find_element_by_link_text("SIM Various Labor Report").click()
        nowhandle=driver.current_window_handle
        allhandles=driver.window_handles

        for handle in allhandles:
            if handle != nowhandle:
                driver.switch_to.window(handle)
                driver.find_element_by_id("CAMUsername").clear()
                driver.find_element_by_id("CAMUsername").send_keys(self.username)
                driver.find_element_by_id("CAMPassword").clear()
                driver.find_element_by_id("CAMPassword").send_keys(self.password)
                driver.find_element_by_id("cmdOK").click()
                time.sleep(3)

                # check for select report level
                driver.find_element_by_css_selector("option[value=\""+report_level_dict[self.record["report_level"]]+"\"]").click()

                #check for select country or company 
                driver.find_element_by_css_selector("option[value=\""+country_or_company_dict[self.record["country_or_company"]]+"\"]").click()

                #check for input date 
                input_date = driver.find_elements_by_css_selector("input[id^=\"txtDateN\"]")
                from_date = '-'.join([str(self.record['start_date'].year), \
                    str(self.record['start_date'].month), str(self.record['start_date'].day)])
                input_date[0].clear()
                input_date[0].send_keys(from_date)

                to_date = '-'.join([str(self.record['end_date'].year), \
                    str(self.record['end_date'].month), str(self.record['end_date'].day)])
                input_date[1].clear()
                input_date[1].send_keys(to_date)

                #check for select report format
                driver.find_element_by_css_selector("option[value=\""+report_format_dict[self.record["report_format"]]+"\"]").click()
                time.sleep(1)

                #check for report criteria 
                driver.find_element_by_css_selector("option[value=\""+report_criteria_dict[self.record["report_criteria"]]+"\"]").click()

                #check for employee or account id
                driver.find_element_by_css_selector("input[value=\""+report_output_level_dict[self.record["report_output_level"]]+"\"]").click()

                if self.record["report_output_level"] == 'Account':
                    driver.find_element_by_css_selector("input[id ^=\"PRMT_TB_\"]").send_keys(self.record["account_id"])
                elif self.record["report_output_level"] == 'Employee':
                    driver.find_element_by_css_selector("input[id ^=\"PRMT_TB_\"]").send_keys(self.record["serial_number"])
                else:
                    raise

                driver.find_element_by_css_selector("button[id ^=\"finishN\"]").click()
                time.sleep(5)
                driver.find_element_by_css_selector("span[lid=\"HLExportReportResult_NS_\"]").click()
                time.sleep(10)
                driver.quit()

    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def tearDown(self):
        self.driver.quit()

def readxls(xlsname):
    try:
        wb = load_workbook(xlsname)
        sheet = wb.get_sheet_by_name("input")
        sheet_parameter = wb.get_sheet_by_name("parameter")
    except Exception as e:
        raise

    start_row = 2
    sheet_rows = []

    for rows in range(start_row, sheet.max_row+1):
        single_row = []
        for cells in sheet[rows]:
            single_row.append(cells.value)
        sheet_rows.append(dict(zip(FIELDS, single_row)))

    #process parameters 
    # check for select report level in column A & B
    global report_level_dict
    report_level_value = []
    report_level_key = []

    for cell in sheet_parameter['A']:
        if cell.value is not None:
            report_level_value.append(cell.value)

    for cell in sheet_parameter['B']:
        if cell.value is not None:
            report_level_key.append(cell.value)

    report_level_dict = dict(zip(report_level_key, report_level_value))

    #check for select country or company 
    global country_or_company_dict 
    country_or_company_value = []
    country_or_company_key = []
    for cell in sheet_parameter['C']:
        if cell.value is not None:
            country_or_company_value.append(cell.value)

    for cell in sheet_parameter['D']:
        if cell.value is not None:
            country_or_company_key.append(cell.value)

    country_or_company_dict = dict(zip(country_or_company_key, country_or_company_value))

    #check for select report format
    global report_format_dict
    report_format_value = []
    report_format_key = []
    for cell in sheet_parameter['E']:
        if cell.value is not None:
            report_format_value.append(cell.value)

    for cell in sheet_parameter['F']:
        if cell.value is not None:
            report_format_key.append(cell.value)

    report_format_dict = dict(zip(report_format_key, report_format_value))

    #check for report criteria 
    global report_criteria_dict
    report_criteria_value = []
    report_criteria_key = []
    for cell in sheet_parameter['G']:
        if cell.value is not None:
            report_criteria_value.append(cell.value)

    for cell in sheet_parameter['H']:
        if cell.value is not None:
            report_criteria_key.append(cell.value)

    report_criteria_dict = dict(zip(report_criteria_key, report_criteria_value))

    #check for report criteria 
    global report_output_level_dict
    report_output_level_value = []
    report_output_level_key = []
    for cell in sheet_parameter['I']:
        if cell.value is not None:
            report_output_level_value.append(cell.value)

    for cell in sheet_parameter['J']:
        if cell.value is not None:
            report_output_level_key.append(cell.value)

    report_output_level_dict = dict(zip(report_output_level_key, report_output_level_value))

    return sheet_rows

def login():
    username = input("Please input your username: ")
    password = getpass("Please entry your password:")

    return username, password

if __name__ == "__main__":
    # predefine username & password for testing purpose
    records = readxls('dlr.xlsx')
    
    username, password = login()
    
    for record in records:
        download_instance = DownloadReport(username, password, **record)
        download_instance.start()

    print('download completed')